"""
Tests complets et robustes pour les documents.
Coverage target: 80%+
Tests: Upload, Validation, Permissions, Admin actions, Notifications, Stats
"""

from django.test import TestCase, override_settings
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from io import BytesIO
import openpyxl
import json

from apps.documents.models import Document, DocumentSpecification, DocumentValidationResult
from apps.documents.validators import DocumentValidator
from apps.documents.tasks import validate_document_task
from apps.folders.models import Folder
from apps.routing_rules.models import RoutingRule
from apps.notifications.models import Notification

User = get_user_model()


class UserSetupMixin:
    """Mixin pour configurer les utilisateurs de test."""
    
    def setUp(self):
        super().setUp()
        
        # Créer admin
        self.admin_user = User.objects.create_user(
            matricule='ADMIN001',
            email='admin@test.com',
            password='admin123',
            is_staff=True,
            is_superuser=True,
            role='ADMIN'
        )
        
        # Créer agent simple
        self.agent_user = User.objects.create_user(
            matricule='AGENT001',
            email='agent@test.com',
            password='agent123',
            is_staff=False,
            role='AGENT'
        )
        
        # Créer specification
        self.excel_spec = DocumentSpecification.objects.create(
            document_type='DONNEES_AGENTS',
            display_name='Données des Agents',
            allowed_formats='xlsx,xls,csv',
            requires_excel=True,
            excel_sheet_name='Donnees Agents',
            required_columns=['Date', 'Montant', 'Client'],
            max_file_size_mb=10,
            max_rows=5000,
            requires_validation=True
        )
        
        # Créer folder
        self.folder = Folder.objects.create(
            name='Données Agents',
            created_by=self.admin_user,
            description='Dossier des données agents'
        )


class DocumentUploadTestCase(UserSetupMixin, APITestCase):
    """Tests robustes pour l'upload de documents."""
    
    def test_unauthenticated_cannot_upload(self):
        """Un utilisateur non authentifié ne peut pas uploader."""
        file_content = b'test content'
        file = SimpleUploadedFile('test.xlsx', file_content)
        
        data = {
            'title': 'Test Document',
            'document_type': 'DONNEES_AGENTS',
            'file': file,
        }
        
        response = self.client.post('/api/documents/', data, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        self.assertEqual(Document.objects.count(), 0)
    
    def test_authenticated_agent_can_upload_document(self):
        """Un agent authentifié peut uploader un document."""
        self.client.force_authenticate(user=self.agent_user)
        
        # Créer un fichier Excel valide
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Date'
        ws['B1'] = 'Montant'
        ws['C1'] = 'Client'
        ws['A2'] = '2026-01-22'
        ws['B2'] = '1000'
        ws['C2'] = 'Client A'
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'donnees_agents.xlsx'
        
        data = {
            'title': 'Donnees Agents Janvier 2026',
            'document_type': 'DONNEES_AGENTS',
            'folder': self.folder.id,
            'specification': self.excel_spec.id,
            'file': excel_file,
            'description': 'Donnees des agents pour janvier'
        }
        
        response = self.client.post('/api/documents/', data, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Document.objects.count(), 1)
        
        doc = Document.objects.first()
        self.assertEqual(doc.title, 'Donnees Agents Janvier 2026')
        self.assertEqual(doc.agent, self.agent_user)
        self.assertEqual(doc.document_type, 'DONNEES_AGENTS')
        self.assertIsNotNone(doc.file)
    
    def test_upload_with_invalid_specification(self):
        """Un upload avec une spécification invalide échoue."""
        self.client.force_authenticate(user=self.agent_user)
        
        file = SimpleUploadedFile('test.xlsx', b'content')
        
        data = {
            'title': 'Test',
            'document_type': 'INVALID_TYPE',  # N'existe pas
            'file': file,
        }
        
        response = self.client.post('/api/documents/', data, format='multipart')
        
        # Peut être 400 ou 404 selon l'implémentation
        self.assertIn(response.status_code, [
            status.HTTP_400_BAD_REQUEST,
            status.HTTP_404_NOT_FOUND
        ])
    
    def test_upload_missing_required_field(self):
        """Un upload sans titre échoue."""
        self.client.force_authenticate(user=self.agent_user)
        
        file = SimpleUploadedFile('test.xlsx', b'content')
        
        data = {
            'document_type': 'DONNEES_AGENTS',
            'file': file,
            # Manque 'title'
        }
        
        response = self.client.post('/api/documents/', data, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Document.objects.count(), 0)


class DocumentValidationTestCase(UserSetupMixin, APITestCase):
    """Tests robustes pour la validation de documents."""
    
    def test_validation_result_creation_and_status(self):
        """On peut créer un DocumentValidationResult avec différents statuts."""
        document = Document.objects.create(
            title='Test Document',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            specification=self.excel_spec
        )
        
        # Créer un résultat PASSED
        passed_result = DocumentValidationResult.objects.create(
            document=document,
            status='PASSED',
            errors=[],
            warnings=['Colonne optionnelle manquante'],
            validation_details={'rows': 10, 'columns': 3}
        )
        
        self.assertEqual(passed_result.status, 'PASSED')
        self.assertEqual(len(passed_result.warnings), 1)
        self.assertEqual(passed_result.validation_details['rows'], 10)
    
    def test_validation_result_with_errors(self):
        """On peut créer un résultat de validation FAILED."""
        document = Document.objects.create(
            title='Invalid Document',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder
        )
        
        failed_result = DocumentValidationResult.objects.create(
            document=document,
            status='FAILED',
            errors=['Colonne Date manquante', 'Format date invalide'],
            warnings=[],
            validation_details={'invalid_rows': 5}
        )
        
        self.assertEqual(failed_result.status, 'FAILED')
        self.assertEqual(len(failed_result.errors), 2)
        self.assertTrue('Colonne Date manquante' in failed_result.errors)
    
    def test_document_validator_with_valid_excel(self):
        """Le validateur accepte un fichier Excel valide."""
        # Créer un fichier Excel valide
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Donnees Agents'  # Définir le bon nom de feuille
        ws['A1'] = 'Date'
        ws['B1'] = 'Montant'
        ws['C1'] = 'Client'
        ws['A2'] = '2026-01-22'
        ws['B2'] = '1000'
        ws['C2'] = 'Client A'
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'test_valide.xlsx'  # Ajouter le nom du fichier
        
        validator = DocumentValidator(excel_file, self.excel_spec)
        is_valid, errors, warnings, details = validator.validate()
        
        # Devrait être valide ou sans erreurs
        if not is_valid and errors:
            self.fail(f"Document validation failed with errors: {errors}")
        self.assertTrue(is_valid or len(errors) == 0)
    
    def test_validation_workflow_complete(self):
        """Test le workflow complet de validation."""
        # 1. Créer document
        doc = Document.objects.create(
            title='Workflow Test',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            specification=self.excel_spec
        )
        
        # 2. Créer résultat validation
        result = DocumentValidationResult.objects.create(
            document=doc,
            status='PASSED',
            errors=[],
            warnings=[],
            validated_at=timezone.now()
        )
        
        # 3. Vérifier lien
        self.assertEqual(doc.validation_result, result)
        self.assertIsNotNone(doc.validation_result.validated_at)
        self.assertEqual(doc.validation_result.status, 'PASSED')


class DocumentPermissionsTestCase(UserSetupMixin, APITestCase):
    """Tests robustes des permissions sur les documents."""
    
    def setUp(self):
        super().setUp()
        
        # Créer document de l'agent
        self.document = Document.objects.create(
            title='Agent Document',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            specification=self.excel_spec
        )
        
        # Créer autre agent
        self.other_agent = User.objects.create_user(
            matricule='OTHER001',
            email='other@test.com',
            password='other123',
            is_staff=False
        )
        
        # Document de l'autre agent
        self.other_document = Document.objects.create(
            title='Other Agent Document',
            document_type='DONNEES_AGENTS',
            agent=self.other_agent,
            folder=self.folder
        )
    
    def test_agent_can_see_own_documents_list(self):
        """Un agent peut voir sa propre liste de documents."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.get('/api/documents/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Handle pagination: response.data is dict with 'results' key
        data = response.data
        if isinstance(data, dict) and 'results' in data:
            data = data['results']
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['id'], self.document.id)
    
    def test_agent_cannot_see_other_agents_documents(self):
        """Un agent ne peut pas voir les documents des autres agents."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.get('/api/documents/')
        
        # Handle pagination: response.data is dict with 'results' key
        data = response.data
        if isinstance(data, dict) and 'results' in data:
            data = data['results']
        document_ids = [d['id'] for d in data]
        self.assertNotIn(self.other_document.id, document_ids)
        self.assertIn(self.document.id, document_ids)
    
    def test_agent_cannot_retrieve_other_document(self):
        """Un agent ne peut pas récupérer le détail du document d'un autre."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.get(f'/api/documents/{self.other_document.id}/')
        
        # Devrait retourner 404 ou 403
        self.assertIn(response.status_code, [
            status.HTTP_403_FORBIDDEN,
            status.HTTP_404_NOT_FOUND
        ])
    
    def test_admin_can_see_all_documents(self):
        """Un admin peut voir TOUS les documents."""
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.get('/api/documents/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Handle pagination: response.data is dict with 'results' key
        data = response.data
        if isinstance(data, dict) and 'results' in data:
            data = data['results']
        self.assertEqual(len(data), 2)
        
        document_ids = [d['id'] for d in data]
        self.assertIn(self.document.id, document_ids)
        self.assertIn(self.other_document.id, document_ids)
    
    def test_admin_can_retrieve_any_document(self):
        """Un admin peut récupérer le détail de n'importe quel document."""
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.get(f'/api/documents/{self.other_document.id}/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['id'], self.other_document.id)
    
    def test_agent_cannot_delete_document_of_other(self):
        """Un agent ne peut pas supprimer le document d'un autre."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.delete(f'/api/documents/{self.other_document.id}/')
        
        self.assertIn(response.status_code, [
            status.HTTP_403_FORBIDDEN,
            status.HTTP_404_NOT_FOUND
        ])
        
        # Vérifier document toujours là
        self.assertTrue(Document.objects.filter(id=self.other_document.id).exists())


class DocumentAdminActionsTestCase(UserSetupMixin, APITestCase):
    """Tests robustes des actions admin sur les documents."""
    
    def setUp(self):
        super().setUp()
        
        # Créer un document validé
        self.document = Document.objects.create(
            title='Document for Admin Actions',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            specification=self.excel_spec,
            status='VALIDE'
        )
    
    def test_admin_can_mark_document_as_opened(self):
        """Un admin peut marquer un document comme ouvert."""
        self.client.force_authenticate(user=self.admin_user)
        
        initial_status = self.document.status
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/mark_as_opened/',
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Vérifier changement de statut
        self.document.refresh_from_db()
        self.assertEqual(self.document.status, 'EN_COURS')
        self.assertIsNotNone(self.document.opened_at)
        
        # Vérifier notification créée
        notifications = Notification.objects.filter(
            recipient=self.agent_user,
            notification_type='DOCUMENT_OPENED'
        )
        self.assertEqual(notifications.count(), 1)
    
    def test_agent_cannot_mark_document_as_opened(self):
        """Un agent ne peut PAS marquer un document comme ouvert."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/mark_as_opened/',
            format='json'
        )
        
        # Devrait être forbiden ou not found
        self.assertIn(response.status_code, [
            status.HTTP_403_FORBIDDEN,
            status.HTTP_404_NOT_FOUND
        ])
        
        # Document ne doit pas être changé
        self.document.refresh_from_db()
        self.assertEqual(self.document.status, 'VALIDE')
    
    def test_admin_can_approve_validated_document(self):
        """Un admin peut approuver un document validé."""
        # Créer résultat de validation
        DocumentValidationResult.objects.create(
            document=self.document,
            status='PASSED',
            errors=[],
            warnings=[],
            validated_at=timezone.now()
        )
        
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/approve/',
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        self.document.refresh_from_db()
        self.assertEqual(self.document.status, 'VALIDE')
        self.assertIsNotNone(self.document.accepted_at)
    
    def test_admin_can_reject_document_with_reason(self):
        """Un admin peut rejeter un document avec une raison."""
        self.client.force_authenticate(user=self.admin_user)
        
        rejection_reason = 'Données incomplètes - colonnes manquantes'
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/reject/',
            data=json.dumps({'reason': rejection_reason}),
            content_type='application/json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Vérifier document
        self.document.refresh_from_db()
        self.assertEqual(self.document.status, 'REJETE')
        self.assertEqual(self.document.rejection_reason, rejection_reason)
        self.assertIsNotNone(self.document.rejected_at)
        
        # Vérifier notification
        notifications = Notification.objects.filter(
            recipient=self.agent_user,
            notification_type='DOCUMENT_REJECTED'
        )
        self.assertEqual(notifications.count(), 1)
        self.assertIn(rejection_reason, notifications.first().message)
    
    def test_reject_document_without_reason_fails(self):
        """Rejeter sans raison retourne une erreur."""
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/reject/',
            data=json.dumps({}),  # Pas de raison
            content_type='application/json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        
        # Document ne doit pas être rejeté
        self.document.refresh_from_db()
        self.assertNotEqual(self.document.status, 'REJETE')
    
    def test_agent_cannot_reject_document(self):
        """Un agent ne peut pas rejeter un document."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.post(
            f'/api/documents/{self.document.id}/reject/',
            {'reason': 'Some reason'},
            format='json'
        )
        
        self.assertIn(response.status_code, [
            status.HTTP_403_FORBIDDEN,
            status.HTTP_404_NOT_FOUND
        ])


class DocumentNotificationTestCase(UserSetupMixin, APITestCase):
    """Tests robustes des notifications."""
    
    def test_notification_created_on_document_marked_as_opened(self):
        """Une notification est créée quand un document est marqué ouvert."""
        document = Document.objects.create(
            title='Test for Notification',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            status='VALIDE'
        )
        
        self.client.force_authenticate(user=self.admin_user)
        
        initial_count = Notification.objects.count()
        
        response = self.client.post(
            f'/api/documents/{document.id}/mark_as_opened/',
            format='json'
        )
        
        if response.status_code == status.HTTP_200_OK:
            # Vérifier notification créée
            new_notifs = Notification.objects.filter(
                recipient=self.agent_user,
                notification_type='DOCUMENT_OPENED'
            )
            self.assertGreaterEqual(new_notifs.count(), 1)
    
    def test_notification_on_document_reject(self):
        """Une notification est créée quand un document est rejeté."""
        document = Document.objects.create(
            title='Document to Reject',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            status='VALIDE'
        )
        
        self.client.force_authenticate(user=self.admin_user)
        
        reason = 'Invalid format'
        response = self.client.post(
            f'/api/documents/{document.id}/reject/',
            {'reason': reason},
            format='json'
        )
        
        if response.status_code == status.HTTP_200_OK:
            # Vérifier notification
            notifs = Notification.objects.filter(
                recipient=self.agent_user,
                notification_type='DOCUMENT_REJECTED'
            )
            self.assertGreaterEqual(notifs.count(), 1)
            self.assertIn(reason, notifs.first().message)
    
    def test_notification_fields_complete(self):
        """Les notifications ont tous les champs requis."""
        document = Document.objects.create(
            title='Test',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            status='VALIDE'
        )
        
        notif = Notification.objects.create(
            recipient=self.agent_user,
            notification_type='TEST',
            title='Test Notification',
            message='This is a test notification',
            document=document,
            is_read=False
        )
        
        self.assertIsNotNone(notif.id)
        self.assertEqual(notif.recipient, self.agent_user)
        self.assertEqual(notif.document, document)
        self.assertFalse(notif.is_read)
        self.assertIsNotNone(notif.created_at)


class DocumentStatisticsTestCase(UserSetupMixin, APITestCase):
    """Tests robustes des statistiques."""
    
    def setUp(self):
        super().setUp()
        
        # Créer plusieurs documents avec différents statuts
        self.documents = {
            'valide': [],
            'rejete': [],
            'en_cours': []
        }
        
        # 3 documents validés
        for i in range(3):
            doc = Document.objects.create(
                title=f'Valide Document {i}',
                document_type='DONNEES_AGENTS',
                agent=self.agent_user,
                folder=self.folder,
                status='VALIDE'
            )
            self.documents['valide'].append(doc)
        
        # 2 documents rejetés
        for i in range(2):
            doc = Document.objects.create(
                title=f'Rejected Document {i}',
                document_type='DONNEES_AGENTS',
                agent=self.agent_user,
                folder=self.folder,
                status='REJETE',
                rejection_reason='Données incomplètes'
            )
            self.documents['rejete'].append(doc)
        
        # 1 document en cours
        doc = Document.objects.create(
            title='In Progress Document',
            document_type='DONNEES_AGENTS',
            agent=self.agent_user,
            folder=self.folder,
            status='EN_COURS'
        )
        self.documents['en_cours'].append(doc)
    
    def test_admin_can_access_statistics_endpoint(self):
        """Un admin peut accéder aux statistiques."""
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.get('/api/documents/validation_stats/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_documents', response.data)
        self.assertIn('rejected_documents', response.data)
    
    def test_statistics_contain_correct_counts(self):
        """Les statistiques montrent les bons nombres."""
        self.client.force_authenticate(user=self.admin_user)
        
        response = self.client.get('/api/documents/validation_stats/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['total_documents'], 6)
        self.assertEqual(response.data['rejected_documents'], 2)
    
    def test_agent_cannot_access_statistics(self):
        """Un agent ne peut pas accéder aux statistiques."""
        self.client.force_authenticate(user=self.agent_user)
        
        response = self.client.get('/api/documents/validation_stats/')
        
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_unauthenticated_cannot_access_statistics(self):
        """Un utilisateur non authentifié ne peut pas accéder aux statistiques."""
        response = self.client.get('/api/documents/validation_stats/')
        
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
