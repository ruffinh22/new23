"""
Checkers pour vérifier si les documents peuvent être envoyés.
Valide l'état, les permissions et les conditions avant d'autoriser un envoi.
"""

from typing import Dict, List, Tuple
from .validators import ValidationService
from .models import Document


class DocumentChecker:
    """
    Classe pour vérifier si un document peut être envoyé.
    Lance une série de vérifications avant d'autoriser l'envoi.
    """

    def __init__(self, document: Document):
        self.document = document
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.checks_results: Dict = {}

    def can_send(self) -> Tuple[bool, Dict]:
        """
        Vérifie si le document peut être envoyé.

        Returns:
            Tuple (can_send, check_results)
        """
        self.errors = []
        self.warnings = []
        self.checks_results = {}

        # Série de vérifications
        self._check_document_exists()
        self._check_file_exists()
        self._check_document_status()
        self._check_validation_status()
        self._check_file_integrity()
        self._check_specifications()
        self._check_permissions()

        result = {
            "can_send": len(self.errors) == 0,
            "errors": self.errors,
            "warnings": self.warnings,
            "checks": self.checks_results,
        }

        return len(self.errors) == 0, result

    def _check_document_exists(self):
        """Vérifie que le document existe et n'est pas supprimé."""
        check_name = "document_exists"

        if not self.document or self.document.pk is None:
            self.errors.append("Le document n'existe pas ou a été supprimé.")
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Document inexistant",
            }
            return

        self.checks_results[check_name] = {
            "status": "PASSED",
            "message": "Document existe",
        }

    def _check_file_exists(self):
        """Vérifie que le fichier existe physiquement."""
        check_name = "file_exists"

        if not self.document.file:
            self.errors.append("Aucun fichier n'est associé au document.")
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Fichier manquant",
            }
            return

        try:
            # Vérifier que le fichier est accessible
            self.document.file.open("rb")
            self.document.file.close()

            self.checks_results[check_name] = {
                "status": "PASSED",
                "message": f"Fichier présent ({self.document.file_size_mb} MB)",
                "file_size": self.document.file_size,
                "file_format": self.document.file_format,
            }
        except Exception as e:
            self.errors.append(f"Le fichier n'est pas accessible: {str(e)}")
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Fichier inaccessible",
            }

    def _check_document_status(self):
        """Vérifie que le document est dans un état valide pour envoi."""
        check_name = "document_status"

        valid_statuses = ["EN_ATTENTE", "VALIDE", "EN_COURS"]

        if self.document.status not in valid_statuses:
            self.errors.append(
                f"Le document ne peut pas être envoyé avec le statut '{self.document.status}'. "
                f"Statuts acceptés: {', '.join(valid_statuses)}"
            )
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": f"Statut invalide: {self.document.status}",
                "current_status": self.document.status,
                "required_statuses": valid_statuses,
            }
            return

        self.checks_results[check_name] = {
            "status": "PASSED",
            "message": f"Statut valide: {self.document.status}",
            "current_status": self.document.status,
        }

    def _check_validation_status(self):
        """Vérifie que le document a été validé."""
        check_name = "validation_status"

        if not self.document.is_validated:
            self.errors.append(
                "Le document n'a pas été validé. "
                "Lancez une validation avant d'envoyer le document."
            )
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Document non validé",
            }
            return

        self.checks_results[check_name] = {
            "status": "PASSED",
            "message": "Document validé",
            "validated_at": self.document.validated_at.isoformat()
            if self.document.validated_at
            else None,
        }

    def _check_file_integrity(self):
        """Vérifie l'intégrité du fichier."""
        check_name = "file_integrity"

        try:
            self.document.file.open("rb")
            file_content = self.document.file.read()
            file_size = len(file_content)
            self.document.file.close()

            # Vérifier que la taille correspond
            if file_size != self.document.file_size:
                self.warnings.append(
                    f"La taille du fichier ({file_size} bytes) diffère de la valeur enregistrée "
                    f"({self.document.file_size} bytes)."
                )
                self.checks_results[check_name] = {
                    "status": "WARNING",
                    "message": "Taille incohérente",
                    "recorded_size": self.document.file_size,
                    "actual_size": file_size,
                }
            else:
                self.checks_results[check_name] = {
                    "status": "PASSED",
                    "message": "Fichier intact",
                    "file_size": file_size,
                }

        except Exception as e:
            self.errors.append(
                f"Impossible de vérifier l'intégrité du fichier: {str(e)}"
            )
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Vérification impossible",
            }

    def _check_specifications(self):
        """Vérifie que le document respecte les spécifications."""
        check_name = "specification_compliance"

        if not self.document.specification:
            self.checks_results[check_name] = {
                "status": "WARNING",
                "message": "Pas de spécification associée",
            }
            return

        spec = self.document.specification

        # Vérifier si la validation est requise
        if not spec.requires_validation:
            self.checks_results[check_name] = {
                "status": "PASSED",
                "message": "Validation non requise pour ce type de document",
            }
            return

        # Ré-valider le fichier pour être sûr
        try:
            self.document.file.open("rb")
            is_valid, validation_data = ValidationService.validate_document(
                self.document.file, spec
            )
            self.document.file.close()

            if is_valid:
                self.checks_results[check_name] = {
                    "status": "PASSED",
                    "message": "Conforme aux spécifications",
                    "validation_details": validation_data.get("details", {}),
                }
            else:
                errors = validation_data.get("errors", [])
                for error in errors:
                    self.errors.append(f"Spécification non respectée: {error}")

                self.checks_results[check_name] = {
                    "status": "FAILED",
                    "message": "Non conforme aux spécifications",
                    "errors": errors,
                }

        except Exception as e:
            self.warnings.append(f"Impossible de re-valider le fichier: {str(e)}")
            self.checks_results[check_name] = {
                "status": "WARNING",
                "message": "Re-validation impossible",
            }

    def _check_permissions(self):
        """Vérifie que le document n'a pas de restrictions."""
        check_name = "permissions"

        # Vérifier que le document n'est pas archivé
        if self.document.status == "ARCHIVE":
            self.errors.append("Le document a été archivé et ne peut pas être envoyé.")
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Document archivé",
            }
            return

        # Vérifier que le document n'a pas été rejeté
        if self.document.status == "REJETE":
            self.errors.append(
                f"Le document a été rejeté. Raison: {self.document.rejection_reason}"
            )
            self.checks_results[check_name] = {
                "status": "FAILED",
                "message": "Document rejeté",
                "rejection_reason": self.document.rejection_reason,
            }
            return

        self.checks_results[check_name] = {
            "status": "PASSED",
            "message": "Aucune restriction",
        }


class FileChecker:
    """Classe pour vérifier les propriétés d'un fichier avant création de document."""

    @staticmethod
    def check_file_validity(file, document_type: str = None) -> Dict:
        """
        Vérifie qu'un fichier est valide.

        Returns:
            Dict avec les résultats de vérification
        """
        checks = {
            "file_name": file.name if hasattr(file, "name") else "unknown",
            "file_size": file.size if hasattr(file, "size") else 0,
            "content_type": file.content_type
            if hasattr(file, "content_type")
            else "unknown",
            "is_valid": True,
            "errors": [],
            "warnings": [],
        }

        # Vérifier que le fichier n'est pas vide
        if file.size == 0:
            checks["is_valid"] = False
            checks["errors"].append("Le fichier est vide.")
            return checks

        # Vérifier l'extension
        file_format = str(file.name).split(".")[-1].lower()
        checks["file_format"] = file_format

        # Pour les fichiers Excel, faire une vérification spéciale
        if file_format in ["xlsx", "xls", "xlsm"]:
            try:
                file.seek(0)
                checker = FileChecker._check_excel_file(file)
                checks["excel_check"] = checker
                if not checker["is_valid"]:
                    checks["is_valid"] = False
                    checks["errors"].extend(checker["errors"])
                    checks["warnings"].extend(checker["warnings"])
            except Exception as e:
                checks["is_valid"] = False
                checks["errors"].append(
                    f"Erreur lors de la vérification du fichier Excel: {str(e)}"
                )

        return checks

    @staticmethod
    def _check_excel_file(file) -> Dict:
        """Vérifie les propriétés spécifiques d'un fichier Excel."""
        try:
            import openpyxl

            file.seek(0)
            wb = openpyxl.load_workbook(file, data_only=False)

            check_result = {
                "is_valid": True,
                "errors": [],
                "warnings": [],
                "sheet_count": len(wb.sheetnames),
                "sheets": wb.sheetnames,
            }

            if not wb.sheetnames:
                check_result["is_valid"] = False
                check_result["errors"].append("Le classeur Excel est vide.")
            else:
                # Vérifier la première feuille
                ws = wb.active
                check_result["active_sheet"] = ws.title
                check_result["row_count"] = ws.max_row
                check_result["column_count"] = ws.max_column

                if ws.max_row == 0:
                    check_result["is_valid"] = False
                    check_result["errors"].append("La feuille active est vide.")

            wb.close()
            return check_result

        except Exception as e:
            return {
                "is_valid": False,
                "errors": [f"Erreur lors de la lecture du fichier Excel: {str(e)}"],
                "warnings": [],
            }


class DocumentValidationChecker:
    """Checker pour effectuer la validation complète avant d'envoyer un document."""

    @staticmethod
    def full_check_before_send(document: Document) -> Tuple[bool, Dict]:
        """
        Effectue une vérification complète avant envoi.

        Returns:
            Tuple (can_send, results)
        """
        checker = DocumentChecker(document)
        return checker.can_send()

    @staticmethod
    def quick_validation_check(document: Document) -> Tuple[bool, str]:
        """
        Effectue une vérification rapide du statut.

        Returns:
            Tuple (is_ready, message)
        """
        if document.status == "REJETE":
            return False, f"Document rejeté: {document.rejection_reason}"

        if document.status == "ARCHIVE":
            return False, "Document archivé et ne peut pas être envoyé"

        if not document.is_validated:
            return False, "Document n'a pas été validé"

        if not document.file:
            return False, "Aucun fichier associé au document"

        return True, "Le document est prêt à être envoyé"
