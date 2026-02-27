"""
Tests pour les validateurs de documents.
"""

import pytest
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth import get_user_model
from io import BytesIO
import openpyxl

from apps.documents.models import (
    DocumentSpecification,
)
from apps.documents.validators import DocumentValidator
from apps.documents.services import DocumentService

User = get_user_model()


@pytest.mark.django_db
class TestDocumentValidator(TestCase):
    """Tests pour la validation de documents."""

    def setUp(self):
        """Configuration initiale."""
        self.user = User.objects.create_user(
            matricule="TEST001",
            email="test@example.com",
            first_name="Test",
            last_name="User",
            department="RH",
            password="testpass123",
        )

        self.spec_excel = DocumentSpecification.objects.create(
            document_type="RAPPORT_EXCEL",
            display_name="Rapport Excel",
            allowed_formats="xlsx,xls",
            requires_excel=True,
            excel_sheet_name="Rapport",
            required_columns=["Date", "Agent", "Heures"],
            max_file_size_mb=50,
            max_rows=1000,
            requires_validation=True,
        )

    def create_excel_file(self, rows=10, columns=None):
        """Crée un fichier Excel test."""
        if columns is None:
            columns = ["Date", "Agent", "Heures"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"

        # En-têtes
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        # Données
        for row in range(2, rows + 2):
            for col in range(1, len(columns) + 1):
                ws.cell(row=row, column=col, value=f"Data{row}{col}")

        # Sauvegarder en mémoire
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        return SimpleUploadedFile(
            "test.xlsx",
            file_obj.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_valid_excel_file(self):
        """Test d'un fichier Excel valide."""
        file_obj = self.create_excel_file(rows=100)
        validator = DocumentValidator(file_obj, self.spec_excel)

        is_valid, errors, warnings, details = validator.validate()

        assert is_valid is True
        assert len(errors) == 0
        assert details["file_format"] == "xlsx"
        assert details["row_count"] == 101  # +1 pour l'en-tête

    def test_missing_required_columns(self):
        """Test avec colonnes manquantes."""
        file_obj = self.create_excel_file(columns=["Date", "Agent"])
        validator = DocumentValidator(file_obj, self.spec_excel)

        is_valid, errors, warnings, details = validator.validate()

        assert is_valid is False
        assert any("manquantes" in error or "Colonnes" in error for error in errors)

    def test_file_too_large(self):
        """Test avec fichier trop volumineux."""
        spec = DocumentSpecification.objects.create(
            document_type="SMALL_FILE",
            display_name="Petit fichier",
            allowed_formats="xlsx",
            max_file_size_mb=0.001,  # Très petit pour le test
            requires_validation=True,
        )

        file_obj = self.create_excel_file(rows=1000)
        validator = DocumentValidator(file_obj, spec)

        is_valid, errors, warnings, details = validator.validate()

        assert is_valid is False
        assert any("taille" in error.lower() for error in errors)

    def test_invalid_file_format(self):
        """Test avec format invalide."""
        spec = DocumentSpecification.objects.create(
            document_type="PDF_ONLY",
            display_name="PDF uniquement",
            allowed_formats="pdf",
            requires_validation=True,
        )

        file_obj = self.create_excel_file()
        validator = DocumentValidator(file_obj, spec)

        is_valid, errors, warnings, details = validator.validate()

        assert is_valid is False
        assert any("format" in error.lower() for error in errors)

    def test_csv_file_validation(self):
        """Test la validation d'un fichier CSV."""
        # Créer un fichier CSV
        csv_content = "Name,Age,City\nJohn,25,NYC\nJane,30,LA\n"
        file_obj = SimpleUploadedFile(
            "test.csv", csv_content.encode("utf-8"), content_type="text/csv"
        )

        spec = DocumentSpecification.objects.create(
            document_type="CSV_FILE",
            display_name="Fichier CSV",
            allowed_formats="csv",
            required_columns=["Name", "Age"],
            max_rows=1000,
            requires_validation=True,
        )

        validator = DocumentValidator(file_obj, spec)
        is_valid, errors, warnings, details = validator.validate()

        assert is_valid is True
        assert details["headers"] == ["Name", "Age", "City"]
        assert details["row_count"] == 2


@pytest.mark.django_db
class TestDocumentService(TestCase):
    """Tests pour le service de documents."""

    def setUp(self):
        """Configuration initiale."""
        self.user = User.objects.create_user(
            matricule="TEST001",
            email="test@example.com",
            first_name="Test",
            last_name="User",
            department="RH",
            password="testpass123",
        )

        self.spec = DocumentSpecification.objects.create(
            document_type="TEST_DOC",
            display_name="Document Test",
            allowed_formats="xlsx",
            max_file_size_mb=50,
            requires_validation=True,
        )

    def create_excel_file(self):
        """Crée un fichier Excel test."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Test"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        return SimpleUploadedFile(
            "test.xlsx",
            file_obj.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_create_document_with_validation(self):
        """Test la création d'un document avec validation."""
        file_obj = self.create_excel_file()

        document, validation_result, is_valid = (
            DocumentService.create_document_with_validation(
                title="Test Document",
                file=file_obj,
                document_type="TEST_DOC",
                agent=self.user,
                auto_validate=True,
            )
        )

        assert document is not None
        assert document.title == "Test Document"
        assert document.is_validated is True
        assert document.status in ["EN_ATTENTE", "VALIDE"]
        assert validation_result is not None

    def test_approve_document(self):
        """Test l'approbation d'un document."""
        file_obj = self.create_excel_file()

        document, _, _ = DocumentService.create_document_with_validation(
            title="Test Document",
            file=file_obj,
            document_type="TEST_DOC",
            agent=self.user,
            auto_validate=True,
        )

        # Approuver
        admin_user = User.objects.create_user(
            matricule="ADMIN001",
            email="admin@example.com",
            password="adminpass",
            first_name="Admin",
            last_name="User",
            department="RH",
            is_staff=True,
        )

        DocumentService.approve_document(document, admin_user)

        document.refresh_from_db()
        assert document.status == "VALIDE"
        assert document.validated_by == admin_user
        assert document.validated_at is not None

    def test_reject_document(self):
        """Test le rejet d'un document."""
        file_obj = self.create_excel_file()

        document, _, _ = DocumentService.create_document_with_validation(
            title="Test Document",
            file=file_obj,
            document_type="TEST_DOC",
            agent=self.user,
            auto_validate=True,
        )

        admin_user = User.objects.create_user(
            matricule="ADMIN001",
            email="admin@example.com",
            password="adminpass",
            first_name="Admin",
            last_name="User",
            department="RH",
            is_staff=True,
        )

        DocumentService.reject_document(document, "Format invalide", admin_user)

        document.refresh_from_db()
        assert document.status == "REJETE"
        assert document.is_validated is False
        assert document.rejection_reason == "Format invalide"


if __name__ == "__main__":
    pytest.main([__file__])
