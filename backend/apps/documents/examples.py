"""
Exemple d'utilisation du système de validation de documents.

Ce fichier montre comment utiliser le système de validation
pour créer et valider des documents.
"""

from django.contrib.auth import get_user_model
from apps.documents.models import Document, DocumentSpecification
from apps.documents.services import DocumentService
from apps.documents.validators import DocumentValidator
from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
from io import BytesIO

User = get_user_model()


# Example 1: Créer une spécification pour un rapport Excel
# ===========================================================
def create_payroll_specification():
    """Crée une spécification pour l'export de paie."""
    spec, created = DocumentSpecification.objects.get_or_create(
        document_type='PAYROLL_REPORT',
        defaults={
            'display_name': 'Rapport de paie',
            'description': 'Export mensuel de la paie des agents',
            'allowed_formats': 'xlsx,xls',
            'requires_excel': True,
            'excel_sheet_name': 'Paie',
            'required_columns': [
                'Date Début',
                'Date Fin',
                'Agent',
                'Matricule',
                'Salaire Brut',
                'Retenues',
                'Salaire Net'
            ],
            'max_file_size_mb': 50,
            'max_rows': 5000,
            'requires_validation': True,
        }
    )
    
    if created:
        print(f"✓ Spécification créée: {spec.display_name}")
    else:
        print(f"✓ Spécification existante: {spec.display_name}")
    
    return spec


# Example 2: Créer un fichier Excel de test
# ==========================================
def create_sample_payroll_file():
    """Crée un fichier Excel de paie valide."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Paie'
    
    # En-têtes
    headers = [
        'Date Début', 'Date Fin', 'Agent', 'Matricule',
        'Salaire Brut', 'Retenues', 'Salaire Net'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données
    data = [
        ['2024-01-01', '2024-01-31', 'Jean Dupont', 'MAT001', 3000, 500, 2500],
        ['2024-01-01', '2024-01-31', 'Marie Martin', 'MAT002', 3500, 600, 2900],
        ['2024-01-01', '2024-01-31', 'Pierre Durand', 'MAT003', 2500, 400, 2100],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Sauvegarder en mémoire
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    
    return SimpleUploadedFile(
        'payroll_january_2024.xlsx',
        file_obj.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# Example 3: Valider un fichier Excel
# ====================================
def validate_payroll_file():
    """Valide un fichier Excel de paie."""
    spec = DocumentSpecification.objects.get(document_type='PAYROLL_REPORT')
    file_obj = create_sample_payroll_file()
    
    print("\n📋 Validation du fichier Excel...")
    
    validator = DocumentValidator(file_obj, spec)
    is_valid, errors, warnings, details = validator.validate()
    
    if is_valid:
        print("✓ Fichier valide!")
        print(f"  Format: {details.get('file_format')}")
        print(f"  Feuille: {details.get('active_sheet')}")
        print(f"  Lignes: {details.get('row_count')}")
        print(f"  Colonnes: {details.get('column_count')}")
        print(f"  En-têtes: {details.get('headers')}")
    else:
        print("✗ Fichier invalide!")
        for error in errors:
            print(f"  ✗ {error}")
    
    if warnings:
        print("⚠ Avertissements:")
        for warning in warnings:
            print(f"  ⚠ {warning}")
    
    return is_valid


# Example 4: Créer un document avec validation automatique
# =========================================================
def upload_document_with_validation():
    """Crée un document et lance la validation."""
    # Récupérer ou créer l'utilisateur
    user, _ = User.objects.get_or_create(
        username='john_doe',
        defaults={
            'email': 'john@example.com',
            'first_name': 'John',
            'last_name': 'Doe',
        }
    )
    
    print("\n📤 Création du document avec validation...")
    
    file_obj = create_sample_payroll_file()
    
    # Créer le document avec validation automatique
    document, validation_result, is_valid = DocumentService.create_document_with_validation(
        title='Rapport de paie - Janvier 2024',
        file=file_obj,
        document_type='PAYROLL_REPORT',
        agent=user,
        description='Export mensuel de la paie pour janvier 2024',
        auto_validate=True
    )
    
    print(f"✓ Document créé: {document.title}")
    print(f"  ID: {document.id}")
    print(f"  Statut: {document.get_status_display()}")
    print(f"  Validé: {'✓ Oui' if document.is_validated else '✗ Non'}")
    
    if validation_result:
        print(f"  Résultat validation: {validation_result.get_status_display()}")
        if validation_result.errors:
            print("  Erreurs:")
            for error in validation_result.errors:
                print(f"    - {error}")
    
    return document


# Example 5: Approuver un document
# ================================
def approve_document(document_id):
    """Approuve un document."""
    document = Document.objects.get(id=document_id)
    
    # Créer ou récupérer un administrateur
    admin, _ = User.objects.get_or_create(
        username='admin',
        defaults={
            'email': 'admin@example.com',
            'is_staff': True,
            'is_superuser': True,
        }
    )
    
    print(f"\n✅ Approbation du document: {document.title}")
    
    if not document.is_validated:
        print("✗ Le document doit être validé avant approbation!")
        return False
    
    DocumentService.approve_document(document, admin)
    
    print(f"✓ Document approuvé!")
    print(f"  Statut: {document.get_status_display()}")
    print(f"  Approuvé par: {document.validated_by.username}")
    
    return True


# Example 6: Rejeter un document
# ==============================
def reject_document(document_id, reason):
    """Rejette un document."""
    document = Document.objects.get(id=document_id)
    
    admin, _ = User.objects.get_or_create(
        username='admin',
        defaults={'email': 'admin@example.com', 'is_staff': True}
    )
    
    print(f"\n❌ Rejet du document: {document.title}")
    
    DocumentService.reject_document(document, reason, admin)
    
    print(f"✓ Document rejeté!")
    print(f"  Raison: {document.rejection_reason}")
    print(f"  Statut: {document.get_status_display()}")
    
    return True


# Example 7: Re-valider un document
# =================================
def revalidate_document(document_id):
    """Re-valide un document."""
    document = Document.objects.get(id=document_id)
    
    print(f"\n🔄 Re-validation du document: {document.title}")
    
    is_valid, validation_result = DocumentService.validate_existing_document(document)
    
    print(f"✓ Re-validation effectuée!")
    print(f"  Valide: {'✓ Oui' if is_valid else '✗ Non'}")
    print(f"  Statut: {document.get_status_display()}")
    
    if validation_result:
        print(f"  Résultat: {validation_result.get_status_display()}")
    
    return is_valid


# Main: Exécuter les exemples
# ============================
if __name__ == '__main__':
    import django
    import os
    
    # Configure Django
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
    django.setup()
    
    print("=" * 60)
    print("EXEMPLES D'UTILISATION DU SYSTÈME DE VALIDATION")
    print("=" * 60)
    
    # 1. Créer la spécification
    create_payroll_specification()
    
    # 2. Valider un fichier
    is_valid = validate_payroll_file()
    
    # 3. Créer et valider un document
    document = upload_document_with_validation()
    
    # 4. Approuver le document
    if is_valid and document.is_validated:
        approve_document(document.id)
    
    # 5. Exemple de rejet (commenté)
    # reject_document(document.id, "Format invalide, données manquantes")
    
    print("\n" + "=" * 60)
    print("✓ Tous les exemples sont terminés!")
    print("=" * 60)
