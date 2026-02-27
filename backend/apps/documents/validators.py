"""
Validateurs pour les documents.
Vérifie les fichiers selon les spécifications définies avant qu'ils soient envoyés.
"""

from typing import Dict, List, Tuple
from django.core.files.uploadedfile import UploadedFile
import openpyxl
import csv
from io import StringIO


# Formats Excel supportés
EXCEL_FORMATS = ["xlsx", "xlsm", "xls", "xlsb", "xlam", "xltx", "xltm", "xlt"]
# Formats texte délimités
DELIMITED_FORMATS = ["csv", "tsv", "txt"]
# Formats OpenDocument
ODS_FORMATS = ["ods"]


class DocumentValidator:
    """Validateur pour vérifier les documents selon leurs spécifications."""

    def __init__(self, document_file: UploadedFile, specification):
        self.file = document_file
        self.specification = specification
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.details: Dict = {}

    def validate(self) -> Tuple[bool, List[str], List[str], Dict]:
        """
        Lance la validation complète du document.

        Returns:
            Tuple (is_valid, errors, warnings, details)
        """
        self.errors = []
        self.warnings = []
        self.details = {}

        # Validation basique
        self._validate_file_exists()
        if self.errors:
            return False, self.errors, self.warnings, self.details

        self._validate_file_size()
        self._validate_file_format()

        if self.errors:
            return False, self.errors, self.warnings, self.details

        # Validation spécifique au type de fichier
        file_format = self._get_file_format()

        # Support de tous les formats Excel
        if file_format in EXCEL_FORMATS:
            self._validate_excel()
        elif file_format in ODS_FORMATS:
            self._validate_ods()
        elif file_format in DELIMITED_FORMATS:
            self._validate_csv()
        elif file_format in ["pdf", "doc", "docx", "txt"]:
            self._validate_document()

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings, self.details

    def _get_file_size(self):
        """Retourne la taille du fichier en gérant BytesIO et les fichiers normaux."""
        if hasattr(self.file, "size"):
            return self.file.size
        elif hasattr(self.file, "seek") and hasattr(self.file, "tell"):
            # BytesIO ou objet file-like
            current_pos = self.file.tell()
            self.file.seek(0, 2)  # Aller à la fin
            file_size = self.file.tell()
            self.file.seek(current_pos)  # Revenir à la position initiale
            return file_size
        return 0

    def _validate_file_exists(self):
        """Vérifie que le fichier existe et n'est pas vide."""
        if not self.file:
            self.errors.append("Le fichier est vide ou manquant.")
            return

        file_size = self._get_file_size()
        if file_size == 0:
            self.errors.append("Le fichier téléchargé est vide.")

    def _validate_file_size(self):
        """Vérifie la taille du fichier."""
        if not self.specification:
            return

        max_size_bytes = self.specification.max_file_size_mb * 1024 * 1024
        file_size = self._get_file_size()

        if file_size > max_size_bytes:
            self.errors.append(
                f"La taille du fichier ({self._format_file_size(file_size)}) "
                f"dépasse la limite maximale ({self.specification.max_file_size_mb} MB)."
            )

    def _validate_file_format(self):
        """Vérifie que le format du fichier est autorisé."""
        if not self.specification:
            return

        file_format = self._get_file_format()
        allowed_formats = self.specification.get_allowed_formats_list()

        if file_format not in allowed_formats:
            self.errors.append(
                f"Format de fichier '{file_format}' non autorisé. "
                f"Formats acceptés: {', '.join(allowed_formats)}"
            )

        self.details["file_format"] = file_format

    def _validate_ods(self):
        """Valide les fichiers ODS (OpenDocument Spreadsheet)."""
        try:
            # ODS est un format ZIP
            import zipfile

            self.file.seek(0)
            if not zipfile.is_zipfile(self.file):
                self.errors.append("Le fichier ODS est corrompu ou invalide.")
                return

            self.details["format_type"] = "ODS (OpenDocument Spreadsheet)"
            self.warnings.append(
                "Format ODS: Support limité, vérifiez la compatibilité Excel."
            )
        except Exception as e:
            self.errors.append(
                f"Erreur lors de la vérification du fichier ODS: {str(e)}"
            )

    def _validate_excel(self):
        """Valide les fichiers Excel (XLSX, XLS, XLSM, XLSB, etc.)."""
        file_format = self._get_file_format()

        # Pour XLSB (Excel Binary), on ne peut faire qu'une validation basique
        if file_format == "xlsb":
            return self._validate_xlsb()

        # Pour XLS (ancien format Excel), essayer avec xlrd si disponible
        if file_format == "xls":
            return self._validate_xls_legacy()

        # Pour XLSX, XLSM, XLT, XLTX, XLTM
        try:
            self.file.seek(0)
            wb = openpyxl.load_workbook(self.file, data_only=True)

            # Vérifier le nom de la feuille si spécifié
            if self.specification.excel_sheet_name:
                if self.specification.excel_sheet_name not in wb.sheetnames:
                    self.errors.append(
                        f"La feuille Excel '{self.specification.excel_sheet_name}' "
                        f"n'existe pas. Feuilles disponibles: {', '.join(wb.sheetnames)}"
                    )
                    return
                ws = wb[self.specification.excel_sheet_name]
            else:
                ws = wb.active

            # Extraire les infos du fichier Excel
            self.details["sheet_names"] = wb.sheetnames
            self.details["active_sheet"] = ws.title
            self.details["row_count"] = ws.max_row
            self.details["column_count"] = ws.max_column
            self.details["format_type"] = f"Excel ({file_format.upper()})"

            # Vérifier le nombre de lignes
            if self.specification.max_rows and ws.max_row > self.specification.max_rows:
                self.errors.append(
                    f"Le fichier contient {ws.max_row} lignes, "
                    f"mais le maximum autorisé est {self.specification.max_rows}."
                )

            # Vérifier les colonnes requises
            required_columns = self.specification.get_required_columns_list()
            if required_columns:
                self._validate_excel_columns(ws, required_columns)

            # Vérifier qu'il y a au moins une ligne de données
            if ws.max_row <= 1:
                self.warnings.append("Le fichier Excel n'a pas de lignes de données.")

            wb.close()

        except openpyxl.utils.exceptions.InvalidFileException:
            self.errors.append("Le fichier Excel est corrompu ou invalide.")
        except Exception as e:
            self.errors.append(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

    def _validate_excel_columns(self, ws, required_columns: List[str]):
        """Valide que les colonnes requises existent dans le fichier Excel."""
        # Récupérer les en-têtes (première ligne)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        self.details["headers"] = headers

        missing_columns = []
        for col in required_columns:
            if col not in headers:
                missing_columns.append(col)

        if missing_columns:
            self.errors.append(
                f"Colonnes manquantes ou invalides: {', '.join(missing_columns)}. "
                f"Colonnes trouvées: {', '.join(headers)}"
            )
        else:
            self.details["required_columns_found"] = True

    def _validate_csv(self):
        """Valide les fichiers CSV."""
        try:
            self.file.seek(0)
            content = self.file.read().decode("utf-8")
            reader = csv.reader(StringIO(content))

            rows = list(reader)
            if not rows:
                self.errors.append("Le fichier CSV est vide.")
                return

            headers = rows[0]
            self.details["headers"] = headers
            self.details["row_count"] = len(rows) - 1  # Exclure l'en-tête
            self.details["column_count"] = len(headers)
            self.details["format_type"] = "CSV (Texte délimité)"

            # Vérifier les colonnes requises
            required_columns = self.specification.get_required_columns_list()
            if required_columns:
                missing_columns = [
                    col for col in required_columns if col not in headers
                ]
                if missing_columns:
                    self.errors.append(
                        f"Colonnes manquantes: {', '.join(missing_columns)}"
                    )
                else:
                    self.details["required_columns_found"] = True

            # Vérifier le nombre de lignes
            if (
                self.specification.max_rows
                and len(rows) - 1 > self.specification.max_rows
            ):
                self.errors.append(
                    f"Le fichier contient {len(rows) - 1} lignes de données, "
                    f"mais le maximum autorisé est {self.specification.max_rows}."
                )

        except UnicodeDecodeError:
            self.errors.append("Le fichier CSV n'est pas encodé en UTF-8.")
        except Exception as e:
            self.errors.append(f"Erreur lors de la lecture du fichier CSV: {str(e)}")

    def _validate_xls_legacy(self):
        """Valide les fichiers XLS (ancien format Excel)."""
        try:
            # Essayer d'abord avec openpyxl (qui peut supporter certains fichiers XLS)
            self.file.seek(0)
            try:
                wb = openpyxl.load_workbook(self.file, data_only=True)
                ws = wb.active
                self.details["sheet_names"] = wb.sheetnames
                self.details["row_count"] = ws.max_row
                self.details["column_count"] = ws.max_column
                self.details["format_type"] = "Excel XLS (ancien format)"
            except (ImportError, KeyError, AttributeError):
                # Si openpyxl échoue, essayer xlrd
                try:
                    import xlrd

                    self.file.seek(0)
                    wb = xlrd.open_workbook(file_contents=self.file.read())
                    ws = wb.sheet_by_index(0)
                    self.details["sheet_names"] = wb.sheet_names()
                    self.details["row_count"] = ws.nrows
                    self.details["column_count"] = ws.ncols
                    self.details["format_type"] = "Excel XLS (ancien format)"
                except ImportError:
                    self.warnings.append(
                        "Format XLS: Support limité (xlrd non installé). "
                        "Veuillez utiliser XLSX à la place."
                    )
                    self.details["format_type"] = "Excel XLS (support limité)"
                    return
        except Exception as e:
            self.errors.append(f"Erreur lors de la lecture du fichier XLS: {str(e)}")

    def _validate_xlsb(self):
        """Valide les fichiers XLSB (Excel Binary)."""
        try:
            import zipfile

            self.file.seek(0)

            # XLSB est un format Microsoft propre, on peut seulement faire une validation basique
            # Vérifier que c'est bien un fichier ZIP (structure XLSB)
            if zipfile.is_zipfile(self.file):
                self.file.seek(0)
                with zipfile.ZipFile(self.file, "r") as zf:
                    # Vérifier la structure XLSB
                    if "xl/workbook.bin" in zf.namelist():
                        self.details["format_type"] = "Excel XLSB (Binaire)"
                        self.warnings.append(
                            "Format XLSB: Validation limitée (lecture binaire). "
                            "Les colonnes ne peuvent pas être vérifiées. "
                            "Veuillez convertir en XLSX pour une validation complète."
                        )
                    else:
                        self.errors.append(
                            "Le fichier XLSB n'a pas la structure attendue."
                        )
            else:
                self.errors.append("Le fichier XLSB est corrompu ou invalide.")
        except ImportError:
            self.errors.append("Le format XLSB nécessite la bibliothèque zipfile.")
        except Exception as e:
            self.errors.append(f"Erreur lors de la lecture du fichier XLSB: {str(e)}")

    def _validate_document(self):
        """Valide les documents (PDF, Word, etc.)."""
        self.file.seek(0)
        content = self.file.read()

        if len(content) < 100:
            self.warnings.append("Le fichier semble très petit, vérifiez-le.")

        self.details["file_size"] = len(content)

    def _get_file_format(self) -> str:
        """Retourne le format du fichier."""
        if not self.file:
            return ""

        # Gérer les objets avec attribut 'name' et les BytesIO
        filename = ""
        if hasattr(self.file, "name"):
            filename = self.file.name
        elif hasattr(self.file, "filename"):
            filename = self.file.filename
        else:
            # Pas de nom, essayer de deviner depuis le type MIME ou retourner vide
            return ""

        return filename.split(".")[-1].lower()

    @staticmethod
    def _format_file_size(bytes_size: int) -> str:
        """Formate la taille d'un fichier en MB."""
        mb = bytes_size / (1024 * 1024)
        return f"{mb:.2f} MB"


class ExcelValidator:
    """Validateur spécifique pour les fichiers Excel."""

    @staticmethod
    def get_file_info(excel_file: UploadedFile) -> Dict:
        """Extrait les informations d'un fichier Excel."""
        info = {}
        try:
            excel_file.seek(0)
            wb = openpyxl.load_workbook(excel_file, data_only=True)

            info["sheet_names"] = wb.sheetnames
            info["active_sheet"] = wb.active.title if wb.active else None
            info["row_count"] = wb.active.max_row if wb.active else 0
            info["column_count"] = wb.active.max_column if wb.active else 0

            # Récupérer les en-têtes
            if wb.active and wb.active.max_row > 0:
                headers = []
                for cell in wb.active[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                info["headers"] = headers

            wb.close()

        except Exception as e:
            info["error"] = str(e)

        return info


class ExcelAdvancedValidator:
    """Validateur avancé pour les fichiers Excel avec vérifications de contenu (XLSX, XLS, XLSM, XLSB, etc.)."""

    def __init__(self, excel_file: UploadedFile, specification=None):
        self.file = excel_file
        self.specification = specification
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.details: Dict = {}
        self.workbook = None
        self.worksheet = None
        self.file_format = str(excel_file.name).split(".")[-1].lower()

    def validate(self) -> Tuple[bool, List[str], List[str], Dict]:
        """
        Valide un fichier Excel de manière avancée.

        Returns:
            Tuple (is_valid, errors, warnings, details)
        """
        self.errors = []
        self.warnings = []
        self.details = {}

        # Gestion spéciale pour XLSB
        if self.file_format == "xlsb":
            return self._validate_xlsb_advanced()

        # Gestion spéciale pour XLS ancien format
        if self.file_format == "xls":
            return self._validate_xls_advanced()

        try:
            # Charger le fichier (XLSX, XLSM, etc.)
            self.file.seek(0)
            self.workbook = openpyxl.load_workbook(self.file, data_only=False)

            # Vérifications basiques
            self._check_workbook_structure()
            if self.errors:
                return False, self.errors, self.warnings, self.details

            # Sélectionner la feuille
            self._select_worksheet()
            if self.errors:
                return False, self.errors, self.warnings, self.details

            # Vérifications de contenu
            self._check_headers()
            self._check_data_rows()
            self._check_data_integrity()
            self._check_formula_errors()

            # Extraire les métadonnées
            self._extract_metadata()

            self.workbook.close()

        except Exception as e:
            self.errors.append(f"Erreur lors du traitement du fichier Excel: {str(e)}")
            return False, self.errors, self.warnings, self.details

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings, self.details

    def _validate_xlsb_advanced(self) -> Tuple[bool, List[str], List[str], Dict]:
        """Valide les fichiers XLSB (Excel Binary) de manière avancée."""
        try:
            import zipfile

            self.file.seek(0)

            if not zipfile.is_zipfile(self.file):
                self.errors.append("Le fichier XLSB est corrompu ou invalide.")
                return False, self.errors, self.warnings, self.details

            self.file.seek(0)
            with zipfile.ZipFile(self.file, "r") as zf:
                if "xl/workbook.bin" not in zf.namelist():
                    self.errors.append("Le fichier XLSB n'a pas la structure attendue.")
                    return False, self.errors, self.warnings, self.details

                self.details["format_type"] = "Excel XLSB (Binaire)"
                self.details["file_size"] = (
                    len(self.file.getvalue())
                    if hasattr(self.file, "getvalue")
                    else self.file.size
                )

                # Pour XLSB, on ne peut pas faire une validation complète des colonnes
                # On avertit l'utilisateur
                self.warnings.append(
                    "Format XLSB: Validation limitée (format binaire propriétaire). "
                    "Les colonnes et données ne peuvent pas être vérifiées en détail. "
                    "Recommandation: Convertissez le fichier en XLSX (.xlsx) pour une validation complète."
                )

        except ImportError:
            self.errors.append(
                "Impossible de valider le format XLSB (zipfile non disponible)."
            )
        except Exception as e:
            self.errors.append(
                f"Erreur lors de la validation du fichier XLSB: {str(e)}"
            )

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings, self.details

    def _validate_xls_advanced(self) -> Tuple[bool, List[str], List[str], Dict]:
        """Valide les fichiers XLS (ancien format Excel) de manière avancée."""
        try:
            self.file.seek(0)

            # Essayer d'abord avec openpyxl
            try:
                self.workbook = openpyxl.load_workbook(self.file, data_only=False)
            except (Exception,):  # Fixed: was bare except
                # Essayer avec xlrd si disponible
                try:
                    import xlrd

                    self.file.seek(0)
                    wb_xlrd = xlrd.open_workbook(file_contents=self.file.read())

                    self.details["format_type"] = "Excel XLS (ancien format, xlrd)"
                    self.details["sheet_names"] = wb_xlrd.sheet_names()
                    self.details["sheet_count"] = len(wb_xlrd.sheet_names())

                    # Récupérer la première feuille
                    sheet = wb_xlrd.sheet_by_index(0)
                    self.details["row_count"] = sheet.nrows
                    self.details["column_count"] = sheet.ncols

                    self.warnings.append(
                        "Format XLS (ancien): Support basique uniquement. "
                        "Recommandation: Convertissez en XLSX (.xlsx) pour une meilleure compatibilité."
                    )

                    is_valid = len(self.errors) == 0
                    return is_valid, self.errors, self.warnings, self.details
                except ImportError:
                    self.warnings.append(
                        "Format XLS: Support limité (xlrd non installé). "
                        "Veuillez installer xlrd ou convertir le fichier en XLSX (.xlsx)."
                    )
                    return True, self.errors, self.warnings, self.details

            # Si openpyxl a fonctionné, continuer avec la validation standard
            self.details["format_type"] = "Excel XLS (ancien format, openpyxl)"
            self._check_workbook_structure()
            if self.errors:
                return False, self.errors, self.warnings, self.details

            self._select_worksheet()
            if self.errors:
                return False, self.errors, self.warnings, self.details

            self._check_headers()
            self._check_data_rows()
            self._check_data_integrity()

            self.workbook.close()

        except Exception as e:
            self.errors.append(f"Erreur lors de la validation du fichier XLS: {str(e)}")
            return False, self.errors, self.warnings, self.details

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings, self.details

    def _check_workbook_structure(self):
        """Vérifie la structure du classeur."""
        if not self.workbook.sheetnames:
            self.errors.append("Le classeur Excel est vide (pas de feuilles).")
            return

        self.details["sheet_names"] = self.workbook.sheetnames
        self.details["sheet_count"] = len(self.workbook.sheetnames)
        self.details["format_type"] = f"Excel ({self.file_format.upper()})"

    def _select_worksheet(self):
        """Sélectionne la feuille appropriée."""
        if self.specification and self.specification.excel_sheet_name:
            if self.specification.excel_sheet_name not in self.workbook.sheetnames:
                self.errors.append(
                    f"La feuille '{self.specification.excel_sheet_name}' n'existe pas. "
                    f"Feuilles disponibles: {', '.join(self.workbook.sheetnames)}"
                )
                return
            self.worksheet = self.workbook[self.specification.excel_sheet_name]
        else:
            self.worksheet = self.workbook.active

        self.details["active_sheet"] = self.worksheet.title if self.worksheet else None

    def _check_headers(self):
        """Vérifie que la première ligne contient des en-têtes valides."""
        if not self.worksheet or self.worksheet.max_row == 0:
            self.errors.append("La feuille Excel est vide.")
            return

        headers = []
        empty_headers = 0

        for cell in self.worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                empty_headers += 1

        if not headers:
            self.errors.append(
                "La première ligne (en-têtes) ne contient aucune valeur."
            )
            return

        if empty_headers > 0:
            self.warnings.append(
                f"La première ligne contient {empty_headers} colonne(s) vide(s)."
            )

        self.details["headers"] = headers
        self.details["header_count"] = len(headers)
        self.details["empty_headers"] = empty_headers

        # Vérifier les colonnes requises
        if self.specification:
            required_columns = self.specification.get_required_columns_list()
            if required_columns:
                missing_columns = [
                    col for col in required_columns if col not in headers
                ]
                if missing_columns:
                    self.errors.append(
                        f"Colonnes requises manquantes: {', '.join(missing_columns)}"
                    )
                else:
                    self.details["required_columns_found"] = True

    def _check_data_rows(self):
        """Vérifie les lignes de données."""
        if not self.worksheet:
            return

        row_count = self.worksheet.max_row - 1  # Exclure l'en-tête

        if row_count == 0:
            self.warnings.append(
                "Le fichier n'a pas de lignes de données (hormis l'en-tête)."
            )
            self.details["data_row_count"] = 0
            return

        self.details["data_row_count"] = row_count
        self.details["total_row_count"] = self.worksheet.max_row
        self.details["column_count"] = self.worksheet.max_column

        # Vérifier le nombre de lignes maximales
        if self.specification and self.specification.max_rows:
            if row_count > self.specification.max_rows:
                self.errors.append(
                    f"Le fichier contient {row_count} lignes de données, "
                    f"mais le maximum autorisé est {self.specification.max_rows}."
                )

        # Détection de lignes vides ou mal formatées
        empty_rows = 0
        partial_rows = 0

        for row_idx in range(2, self.worksheet.max_row + 1):
            row_values = [cell.value for cell in self.worksheet[row_idx]]

            if all(v is None or str(v).strip() == "" for v in row_values):
                empty_rows += 1
            elif sum(1 for v in row_values if v is not None) < len(row_values) / 2:
                partial_rows += 1

        if empty_rows > 0:
            self.warnings.append(
                f"Détectées {empty_rows} ligne(s) vide(s) dans le fichier."
            )

        if partial_rows > 0:
            self.warnings.append(
                f"Détectées {partial_rows} ligne(s) partiellement remplie(s)."
            )

        self.details["empty_rows"] = empty_rows
        self.details["partial_rows"] = partial_rows

    def _check_data_integrity(self):
        """Vérifie l'intégrité des données."""
        if not self.worksheet or self.worksheet.max_row <= 1:
            return

        data_issues = {
            "merged_cells": len(self.worksheet.merged_cells.ranges),
            "hidden_rows": 0,
            "hidden_columns": 0,
            "cells_with_formulas": 0,
        }

        # Compter les formules
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":  # 'f' = formula
                    data_issues["cells_with_formulas"] += 1

        # Compter les lignes/colonnes cachées
        for row in self.worksheet.row_dimensions.values():
            if row.hidden:
                data_issues["hidden_rows"] += 1

        for col in self.worksheet.column_dimensions.values():
            if col.hidden:
                data_issues["hidden_columns"] += 1

        self.details["data_integrity"] = data_issues

        if data_issues["merged_cells"] > 0:
            self.warnings.append(
                f"Le fichier contient {data_issues['merged_cells']} cellule(s) fusionnée(s)."
            )

        if data_issues["hidden_rows"] > 0:
            self.warnings.append(
                f"Le fichier contient {data_issues['hidden_rows']} ligne(s) cachée(s)."
            )

    def _check_formula_errors(self):
        """Vérifie s'il y a des erreurs de formule."""
        if not self.worksheet:
            return

        formula_errors = []

        for row in self.worksheet.iter_rows():
            for cell in row:
                # Vérifier les erreurs Excel
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("#")
                ):
                    formula_errors.append(
                        f"Erreur de formule à {cell.coordinate}: {cell.value}"
                    )

        if formula_errors:
            for error in formula_errors[:5]:  # Limiter à 5 erreurs affichées
                self.errors.append(error)

            if len(formula_errors) > 5:
                self.errors.append(
                    f"... et {len(formula_errors) - 5} autre(s) erreur(s) de formule."
                )

    def _extract_metadata(self):
        """Extrait les métadonnées du fichier."""
        if not self.worksheet:
            return

        # Propriétés du classeur
        if self.workbook.properties:
            props = self.workbook.properties
            self.details["file_metadata"] = {
                "creator": props.creator,
                "created": str(props.created) if props.created else None,
                "modified": str(props.modified) if props.modified else None,
                "title": props.title,
                "subject": props.subject,
            }


class ValidationService:
    """Service pour gérer la validation des documents."""

    @staticmethod
    def validate_document(
        document_file: UploadedFile, specification
    ) -> Tuple[bool, Dict]:
        """
        Valide un document et retourne le résultat.

        Returns:
            Tuple (is_valid, validation_data)
        """
        # Vérifier le type de fichier pour utiliser le validateur approprié
        file_format = str(document_file.name).split(".")[-1].lower()

        if file_format in ["xlsx", "xls", "xlsm"]:
            # Utiliser le validateur Excel avancé pour les fichiers Excel
            validator = ExcelAdvancedValidator(document_file, specification)
        else:
            # Utiliser le validateur standard pour les autres fichiers
            validator = DocumentValidator(document_file, specification)

        is_valid, errors, warnings, details = validator.validate()

        validation_data = {
            "status": "PASSED" if is_valid else ("WARNING" if warnings else "FAILED"),
            "is_valid": is_valid,
            "errors": errors,
            "warnings": warnings,
            "details": details,
        }

        return is_valid, validation_data
