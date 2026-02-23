"""
Module d'extraction de métadonnées de fichiers.

Extrait les informations complètes des fichiers selon leur type:
- Excel: nombre de lignes, colonnes, feuilles
- PDF: nombre de pages
- Images: dimensions (largeur x hauteur)
- CSV/TSV: nombre de lignes et colonnes
"""

import io
import csv
from typing import Dict, Any, Optional
from django.core.files.uploadedfile import UploadedFile
import logging

logger = logging.getLogger(__name__)


class FileMetadataExtractor:
    """Extracteur de métadonnées pour différents types de fichiers."""

    @staticmethod
    def extract_metadata(file: UploadedFile) -> Dict[str, Any]:
        """
        Extrait les métadonnées d'un fichier selon son type.

        Args:
            file: Fichier uploadé

        Returns:
            Dict avec métadonnées extraites
        """
        file_name = file.name.lower()
        metadata = {
            'file_size_kb': file.size / 1024,
            'file_size_mb': file.size / (1024 * 1024),
        }

        try:
            # Déterminer le type de fichier
            if file_name.endswith('.pdf'):
                metadata.update(FileMetadataExtractor._extract_pdf_metadata(file))

            elif file_name.endswith(('.xlsx', '.xlsm', '.xls', '.ods')):
                metadata.update(FileMetadataExtractor._extract_spreadsheet_metadata(file))

            elif file_name.endswith(('.csv', '.tsv')):
                metadata.update(FileMetadataExtractor._extract_csv_metadata(file))

            elif file_name.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp')):
                metadata.update(FileMetadataExtractor._extract_image_metadata(file))

            elif file_name.endswith(('.docx', '.doc')):
                metadata.update(FileMetadataExtractor._extract_document_metadata(file))

            elif file_name.endswith('.zip'):
                metadata.update(FileMetadataExtractor._extract_zip_metadata(file))

        except Exception as e:
            logger.warning(f"Erreur extraction métadonnées {file_name}: {str(e)}")
            metadata['extraction_error'] = str(e)

        return metadata

    @staticmethod
    def _extract_pdf_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées PDF (nombre de pages)."""
        try:
            import PyPDF2
            file.seek(0)
            pdf_reader = PyPDF2.PdfReader(file)
            return {
                'pages': len(pdf_reader.pages),
                'is_encrypted': pdf_reader.is_encrypted,
            }
        except ImportError:
            logger.warning("PyPDF2 non installé, impossible de lire les métadonnées PDF")
            return {'pages': None}
        except Exception as e:
            logger.warning(f"Erreur lecture PDF: {str(e)}")
            return {'pages': None}

    @staticmethod
    def _extract_spreadsheet_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées pour feuilles de calcul (Excel, ODS)."""
        try:
            import openpyxl
            file.seek(0)
            wb = openpyxl.load_workbook(file, data_only=True)

            metadata = {
                'sheets': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
            }

            # Analyser chaque feuille
            max_rows = 0
            max_cols = 0

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                if ws.max_row > max_rows:
                    max_rows = ws.max_row
                if ws.max_column > max_cols:
                    max_cols = ws.max_column

            metadata['rows'] = max_rows
            metadata['columns'] = max_cols

            wb.close()
            return metadata

        except ImportError:
            logger.warning("openpyxl non installé")
            return {'sheets': None, 'rows': None, 'columns': None}
        except Exception as e:
            logger.warning(f"Erreur lecture Excel: {str(e)}")
            return {'sheets': None, 'rows': None, 'columns': None}

    @staticmethod
    def _extract_csv_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées CSV/TSV (lignes et colonnes)."""
        try:
            file.seek(0)
            # Déterminer le délimiteur
            delimiter = '\t' if file.name.lower().endswith('.tsv') else ','

            # Lire le fichier
            text_file = io.TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.reader(text_file, delimiter=delimiter)

            rows = 0
            columns = 0

            for idx, row in enumerate(reader):
                rows = idx + 1
                if idx == 0:
                    columns = len(row)

            file.seek(0)
            return {
                'rows': rows,
                'columns': columns,
            }

        except Exception as e:
            logger.warning(f"Erreur lecture CSV: {str(e)}")
            return {'rows': None, 'columns': None}

    @staticmethod
    def _extract_image_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées images (dimensions)."""
        try:
            from PIL import Image
            file.seek(0)
            img = Image.open(file)
            width, height = img.size

            file.seek(0)
            return {
                'width': width,
                'height': height,
                'format': img.format,
            }

        except ImportError:
            logger.warning("Pillow non installé")
            return {'width': None, 'height': None}
        except Exception as e:
            logger.warning(f"Erreur lecture image: {str(e)}")
            return {'width': None, 'height': None}

    @staticmethod
    def _extract_document_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées documents (Word - pages estimées)."""
        try:
            from docx import Document
            file.seek(0)
            doc = Document(file)

            # Compter paragraphes et sections comme approximation de pages
            paragraph_count = len(doc.paragraphs)
            # Estimer: ~45 lignes par page en moyenne
            estimated_pages = max(1, paragraph_count // 45)

            file.seek(0)
            return {
                'paragraphs': paragraph_count,
                'estimated_pages': estimated_pages,
            }

        except ImportError:
            logger.warning("python-docx non installé")
            return {'paragraphs': None, 'estimated_pages': None}
        except Exception as e:
            logger.warning(f"Erreur lecture DOCX: {str(e)}")
            return {'paragraphs': None, 'estimated_pages': None}

    @staticmethod
    def _extract_zip_metadata(file: UploadedFile) -> Dict[str, Any]:
        """Extrait métadonnées ZIP (nombre de fichiers)."""
        try:
            import zipfile
            file.seek(0)
            with zipfile.ZipFile(file) as zf:
                file_count = len(zf.namelist())
                total_uncompressed = sum(info.file_size for info in zf.infolist())

            file.seek(0)
            return {
                'files': file_count,
                'uncompressed_size_mb': total_uncompressed / (1024 * 1024),
            }

        except Exception as e:
            logger.warning(f"Erreur lecture ZIP: {str(e)}")
            return {'files': None}
